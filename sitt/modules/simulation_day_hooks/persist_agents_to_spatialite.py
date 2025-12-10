# SPDX-FileCopyrightText: 2022-present Maximilian Kalus <info@auxnet.de>
#
# SPDX-License-Identifier: MIT
"""
Persist agents' routes to a Spatialite/GEoPackage database. We will save each day separately, so it is easier to
comprehend the data.
"""
import datetime as dt
import logging
import os
import shutil
import sqlite3
from typing import Iterable

import fiona
import igraph as ig
from shapely import LineString, force_2d, union_all, MultiLineString
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, NamedStyle

from sitt import SimulationDayHookInterface, Configuration, Context, Agent, SetOfResults

logger = logging.getLogger()

class PersistAgentsToSpatialite(SimulationDayHookInterface):
    """
    Persist agents' routes to a Spatialite/GEoPackage database. We will save each day separately, so it is easier to
    comprehend the data.
    """
    def __init__(self, ignore_empty_agents: bool = True, only_unique: bool = True):
        super().__init__()
        self.ignore_empty_agents: bool = ignore_empty_agents
        """Ignore agents with no routes."""
        self.only_unique: bool =  only_unique
        """Save unique routes only (never the same ones)."""
        self.folder: str|None = None
        self.min_time: dt.datetime = dt.datetime.now()
        self.counters: dict[str, dict[str, int]] = {}
        """Counters for routes."""
        self.route_graph: ig.Graph = ig.Graph(directed=True)
        """Keep start and end vertices as graph, so we can find the shortest paths in the end."""

    def _initialize(self, config: Configuration):
        # set min time
        self.min_time = dt.datetime.combine(config.start_date, dt.datetime.min.time())

        # create folder name
        self.folder = f"simulation_{config.simulation_route}_{config.start_date}"

        # remove old data if it exists
        if os.path.exists(self.folder):
            shutil.rmtree(self.folder)

        # create folder
        os.mkdir(self.folder)

        logger.info(f"Saving data to folder {self.folder}")


    def run(self, config: Configuration, context: Context, agents: list[Agent], agents_finished_for_today: list[Agent],
            results: SetOfResults, current_day: int) -> list[Agent]:
        if self.skip:
            return agents_finished_for_today

        # initialize output
        if self.folder is None:
            self._initialize(config)

        self._persist_agents(agents_finished_for_today, context, current_day)

        return agents_finished_for_today

    def finish_simulation(self, results: SetOfResults, config: Configuration, context: Context, current_day: int) -> None:
        self._save_routes(config, context)
        self._calculate_totals(context)

        logger.info(f"Saved data to {self.folder}")

    def _save_routes(self, config: Configuration, context: Context):
        filename = os.path.join(self.folder, "routes.gpkg")
        excel_filename = os.path.join(self.folder, "routes.xlsx")
        logger.info(f"Saving routes to {filename} and {excel_filename}")

        out = fiona.open(filename, 'w', driver='GPKG', crs='EPSG:4326', schema={'geometry': 'MultiLineString', 'properties': {'id': 'str', 'length_hrs': 'float', 'end_hub': 'str', 'end_time': 'datetime', 'start_hubs': 'str', 'start_times': 'str', 'overnight_hubs': 'str'}})

        wb = Workbook()
        if 'header' not in wb.named_styles:
            my_header = NamedStyle(name="header")
            my_header.font.bold = True
            wb.add_named_style(my_header)

        ws = wb.active
        ws.title = "Routes"
        ws.append(['ID', 'Length (hrs)', 'Arrival Day', 'End Hub', 'End Time', 'Start Hubs', 'Start Times', 'Overnight Hubs'])
        for cell in ws['1:1']:
            cell.style = 'header'

        for endpoint in self.route_graph.vs.select(is_finished=True):
            routes = set()
            start_hubs = set()
            start_times = set()
            overnight_hubs = set()
            lowest_time = endpoint['start_time']

            for v in self.route_graph.bfsiter(endpoint):
                for r in v['route']:
                    routes.add(r)
                start_hub = v['start_hub']
                if start_hub in config.simulation_starts:
                    start_hubs.add(v['start_hub'])
                    start_times.add(v['start_time'].strftime('%Y-%m-%d %H:%M'))
                    if v['start_time'] < lowest_time:
                        lowest_time = v['start_time']
                else:
                    overnight_hubs.add(v['start_hub'])

            geom = self._create_route_from_edge_ids(context, routes)

            difference = (endpoint['end_time'] - lowest_time).total_seconds() / 3600  # convert to hours
            diff_padded = f'{difference:.2f}'.rjust(7, '0')
            my_id = f'{diff_padded}_{endpoint["end_hub"]}'
            stat_hubs = ', '.join(list(start_hubs))
            start_times = ', '.join(list(start_times))
            overnight_hubs = ', '.join(list(overnight_hubs))

            out.write({'geometry': geom, 'properties': {
                'id': my_id,
                'length_hrs': difference,
                'end_hub': endpoint['end_hub'],
                'end_time': endpoint['end_time'],
                'start_hubs': stat_hubs,
                'start_times': start_times,
                'overnight_hubs': overnight_hubs,
            }})

            ws.append([my_id, difference, endpoint['day'], endpoint['end_hub'], endpoint['end_time'], stat_hubs, start_times, overnight_hubs])
        out.close()

        # copy styles for QGIS
        con = sqlite3.connect(filename)
        cur = con.cursor()
        cur.execute("CREATE TABLE layer_styles (id INTEGER PRIMARY KEY, f_table_catalog TEXT, f_table_schema TEXT, f_table_name TEXT, f_geometry_column TEXT, styleName TEXT, styleQML TEXT, styleSLD TEXT, useAsDefault BOOLEAN, description TEXT, owner TEXT, ui TEXT, update_time DATETIME)")
        con.commit()
        cur.execute("INSERT INTO layer_styles (id, f_table_catalog, f_table_schema, f_table_name, f_geometry_column, styleName, styleQML, styleSLD, useAsDefault, description, owner, ui, update_time) VALUES (1, '', '', 'routes', 'geom', 'style', '<!DOCTYPE qgis PUBLIC ''http://mrcc.com/qgis.dtd'' ''SYSTEM''> <qgis styleCategories=\"Symbology\" version=\"3.44.5-Solothurn\">  <renderer-v2 referencescale=\"-1\" forceraster=\"0\" type=\"singleSymbol\" enableorderby=\"0\" symbollevels=\"0\">   <symbols>    <symbol frame_rate=\"10\" is_animated=\"0\" name=\"0\" force_rhr=\"0\" type=\"line\" clip_to_extent=\"1\" alpha=\"1\">     <data_defined_properties>      <Option type=\"Map\">       <Option value=\"\" name=\"name\" type=\"QString\"/>       <Option name=\"properties\"/>       <Option value=\"collection\" name=\"type\" type=\"QString\"/>      </Option>     </data_defined_properties>     <layer pass=\"0\" class=\"SimpleLine\" id=\"{214833f4-dcad-4ca6-bb47-114a2110e960}\" locked=\"0\" enabled=\"1\">      <Option type=\"Map\">       <Option value=\"0\" name=\"align_dash_pattern\" type=\"QString\"/>       <Option value=\"square\" name=\"capstyle\" type=\"QString\"/>       <Option value=\"5;2\" name=\"customdash\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"customdash_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"customdash_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"dash_pattern_offset\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"dash_pattern_offset_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"dash_pattern_offset_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"draw_inside_polygon\" type=\"QString\"/>       <Option value=\"bevel\" name=\"joinstyle\" type=\"QString\"/>       <Option value=\"72,123,182,255,rgb:0.2823529,0.4823529,0.7137255,1\" name=\"line_color\" type=\"QString\"/>       <Option value=\"solid\" name=\"line_style\" type=\"QString\"/>       <Option value=\"0.66\" name=\"line_width\" type=\"QString\"/>       <Option value=\"MM\" name=\"line_width_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"offset\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"offset_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"offset_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"ring_filter\" type=\"QString\"/>       <Option value=\"0\" name=\"trim_distance_end\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"trim_distance_end_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"trim_distance_end_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"trim_distance_start\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"trim_distance_start_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"trim_distance_start_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"tweak_dash_pattern_on_corners\" type=\"QString\"/>       <Option value=\"0\" name=\"use_custom_dash\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"width_map_unit_scale\" type=\"QString\"/>      </Option>      <data_defined_properties>       <Option type=\"Map\">        <Option value=\"\" name=\"name\" type=\"QString\"/>        <Option name=\"properties\"/>        <Option value=\"collection\" name=\"type\" type=\"QString\"/>       </Option>      </data_defined_properties>     </layer>    </symbol>   </symbols>   <rotation/>   <sizescale/>   <data-defined-properties>    <Option type=\"Map\">     <Option value=\"\" name=\"name\" type=\"QString\"/>     <Option name=\"properties\"/>     <Option value=\"collection\" name=\"type\" type=\"QString\"/>    </Option>   </data-defined-properties>  </renderer-v2>  <selection mode=\"Default\">   <selectionColor invalid=\"1\"/>   <selectionSymbol>    <symbol frame_rate=\"10\" is_animated=\"0\" name=\"\" force_rhr=\"0\" type=\"line\" clip_to_extent=\"1\" alpha=\"1\">     <data_defined_properties>      <Option type=\"Map\">       <Option value=\"\" name=\"name\" type=\"QString\"/>       <Option name=\"properties\"/>       <Option value=\"collection\" name=\"type\" type=\"QString\"/>      </Option>     </data_defined_properties>     <layer pass=\"0\" class=\"SimpleLine\" id=\"{de8ac958-4167-4765-b8e7-9ceedea1870a}\" locked=\"0\" enabled=\"1\">      <Option type=\"Map\">       <Option value=\"0\" name=\"align_dash_pattern\" type=\"QString\"/>       <Option value=\"square\" name=\"capstyle\" type=\"QString\"/>       <Option value=\"5;2\" name=\"customdash\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"customdash_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"customdash_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"dash_pattern_offset\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"dash_pattern_offset_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"dash_pattern_offset_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"draw_inside_polygon\" type=\"QString\"/>       <Option value=\"bevel\" name=\"joinstyle\" type=\"QString\"/>       <Option value=\"35,35,35,255,rgb:0.1372549,0.1372549,0.1372549,1\" name=\"line_color\" type=\"QString\"/>       <Option value=\"solid\" name=\"line_style\" type=\"QString\"/>       <Option value=\"0.26\" name=\"line_width\" type=\"QString\"/>       <Option value=\"MM\" name=\"line_width_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"offset\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"offset_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"offset_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"ring_filter\" type=\"QString\"/>       <Option value=\"0\" name=\"trim_distance_end\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"trim_distance_end_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"trim_distance_end_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"trim_distance_start\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"trim_distance_start_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"trim_distance_start_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"tweak_dash_pattern_on_corners\" type=\"QString\"/>       <Option value=\"0\" name=\"use_custom_dash\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"width_map_unit_scale\" type=\"QString\"/>      </Option>      <data_defined_properties>       <Option type=\"Map\">        <Option value=\"\" name=\"name\" type=\"QString\"/>        <Option name=\"properties\"/>        <Option value=\"collection\" name=\"type\" type=\"QString\"/>       </Option>      </data_defined_properties>     </layer>    </symbol>   </selectionSymbol>  </selection>  <blendMode>0</blendMode>  <featureBlendMode>0</featureBlendMode>  <layerGeometryType>1</layerGeometryType> </qgis>', '<?xml version= \"1.0 \" encoding= \"UTF-8 \"?> <StyledLayerDescriptor xmlns= \"http://www.opengis.net/sld \" xsi:schemaLocation= \"http://www.opengis.net/sld http://schemas.opengis.net/sld/1.1.0/StyledLayerDescriptor.xsd \" xmlns:ogc= \"http://www.opengis.net/ogc \" xmlns:xsi= \"http://www.w3.org/2001/XMLSchema-instance \" xmlns:xlink= \"http://www.w3.org/1999/xlink \" xmlns:se= \"http://www.opengis.net/se \" version= \"1.1.0 \">  <NamedLayer>   <se:Name>routes</se:Name>   <UserStyle>    <se:Name>routes</se:Name>    <se:FeatureTypeStyle>     <se:Rule>      <se:Name>Single symbol</se:Name>      <se:LineSymbolizer>       <se:Stroke>        <se:SvgParameter name= \"stroke \">#487bb6</se:SvgParameter>        <se:SvgParameter name= \"stroke-width \">2</se:SvgParameter>        <se:SvgParameter name= \"stroke-linejoin \">bevel</se:SvgParameter>        <se:SvgParameter name= \"stroke-linecap \">square</se:SvgParameter>       </se:Stroke>      </se:LineSymbolizer>     </se:Rule>    </se:FeatureTypeStyle>   </UserStyle>  </NamedLayer> </StyledLayerDescriptor>', 0, 'Mi. Dez. 10 12:22:15 2025', '', null, '2025-12-10T11:22:15.000Z');")

        con.commit()
        con.close()

        # now do cancelled routes
        wb.create_sheet(title="Cancelled")
        ws = wb["Cancelled"]
        ws.append(['ID', 'Length (hrs)', 'Arrival Day', 'End Hub', 'End Time', 'Start Hubs', 'Start Times', 'Overnight Hubs'])
        for cell in ws['1:1']:
            cell.style = 'header'

        filename = os.path.join(self.folder, "routes_cancelled.gpkg")
        out = fiona.open(filename, 'w', driver='GPKG', crs='EPSG:4326', schema={'geometry': 'MultiLineString',
                                                                                'properties': {'id': 'str',
                                                                                               'length_hrs': 'float',
                                                                                               'end_hub': 'str',
                                                                                               'end_time': 'datetime',
                                                                                               'start_hubs': 'str',
                                                                                               'start_times': 'str',
                                                                                               'overnight_hubs': 'str'}})

        for endpoint in self.route_graph.vs.select(is_cancelled=True):
            routes = set()
            start_hubs = set()
            start_times = set()
            overnight_hubs = set()
            lowest_time = endpoint['start_time']

            for v in self.route_graph.bfsiter(endpoint):
                for r in v['route']:
                    routes.add(r)
                start_hub = v['start_hub']
                if start_hub in config.simulation_starts:
                    start_hubs.add(v['start_hub'])
                    start_times.add(v['start_time'].strftime('%Y-%m-%d %H:%M'))
                    if v['start_time'] < lowest_time:
                        lowest_time = v['start_time']
                else:
                    overnight_hubs.add(v['start_hub'])

            geom = self._create_route_from_edge_ids(context, routes)

            difference = (endpoint['end_time'] - lowest_time).total_seconds() / 3600  # convert to hours
            diff_padded = f'{difference:.2f}'.rjust(7, '0')
            my_id = f'{diff_padded}_{endpoint["end_hub"]}'
            stat_hubs = ', '.join(list(start_hubs))
            start_times = ', '.join(list(start_times))
            overnight_hubs = ', '.join(list(overnight_hubs))

            out.write({'geometry': geom, 'properties': {
                'id': my_id,
                'length_hrs': difference,
                'end_hub': endpoint['end_hub'],
                'end_time': endpoint['end_time'],
                'start_hubs': stat_hubs,
                'start_times': start_times,
                'overnight_hubs': overnight_hubs,
            }})

            ws.append([my_id, difference, endpoint['day'], endpoint['end_hub'], endpoint['end_time'], stat_hubs, start_times, overnight_hubs])

        out.close()

        # copy styles for QGIS
        con = sqlite3.connect(filename)
        cur = con.cursor()
        cur.execute("CREATE TABLE layer_styles (id INTEGER PRIMARY KEY, f_table_catalog TEXT, f_table_schema TEXT, f_table_name TEXT, f_geometry_column TEXT, styleName TEXT, styleQML TEXT, styleSLD TEXT, useAsDefault BOOLEAN, description TEXT, owner TEXT, ui TEXT, update_time DATETIME)")
        con.commit()
        cur.execute("INSERT INTO layer_styles (id, f_table_catalog, f_table_schema, f_table_name, f_geometry_column, styleName, styleQML, styleSLD, useAsDefault, description, owner, ui, update_time) VALUES (1, '', '', 'routes_cancelled', 'geom', 'style', '<!DOCTYPE qgis PUBLIC ''http://mrcc.com/qgis.dtd'' ''SYSTEM''> <qgis styleCategories=\"Symbology\" version=\"3.44.5-Solothurn\">  <renderer-v2 referencescale=\"-1\" forceraster=\"0\" type=\"singleSymbol\" enableorderby=\"0\" symbollevels=\"0\">   <symbols>    <symbol frame_rate=\"10\" is_animated=\"0\" name=\"0\" force_rhr=\"0\" type=\"line\" clip_to_extent=\"1\" alpha=\"1\">     <data_defined_properties>      <Option type=\"Map\">       <Option value=\"\" name=\"name\" type=\"QString\"/>       <Option name=\"properties\"/>       <Option value=\"collection\" name=\"type\" type=\"QString\"/>      </Option>     </data_defined_properties>     <layer pass=\"0\" class=\"SimpleLine\" id=\"{214833f4-dcad-4ca6-bb47-114a2110e960}\" locked=\"0\" enabled=\"1\">      <Option type=\"Map\">       <Option value=\"0\" name=\"align_dash_pattern\" type=\"QString\"/>       <Option value=\"square\" name=\"capstyle\" type=\"QString\"/>       <Option value=\"5;2\" name=\"customdash\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"customdash_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"customdash_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"dash_pattern_offset\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"dash_pattern_offset_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"dash_pattern_offset_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"draw_inside_polygon\" type=\"QString\"/>       <Option value=\"bevel\" name=\"joinstyle\" type=\"QString\"/>       <Option value=\"72,123,182,255,rgb:0.2823529,0.4823529,0.7137255,1\" name=\"line_color\" type=\"QString\"/>       <Option value=\"solid\" name=\"line_style\" type=\"QString\"/>       <Option value=\"0.66\" name=\"line_width\" type=\"QString\"/>       <Option value=\"MM\" name=\"line_width_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"offset\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"offset_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"offset_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"ring_filter\" type=\"QString\"/>       <Option value=\"0\" name=\"trim_distance_end\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"trim_distance_end_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"trim_distance_end_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"trim_distance_start\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"trim_distance_start_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"trim_distance_start_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"tweak_dash_pattern_on_corners\" type=\"QString\"/>       <Option value=\"0\" name=\"use_custom_dash\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"width_map_unit_scale\" type=\"QString\"/>      </Option>      <data_defined_properties>       <Option type=\"Map\">        <Option value=\"\" name=\"name\" type=\"QString\"/>        <Option name=\"properties\"/>        <Option value=\"collection\" name=\"type\" type=\"QString\"/>       </Option>      </data_defined_properties>     </layer>    </symbol>   </symbols>   <rotation/>   <sizescale/>   <data-defined-properties>    <Option type=\"Map\">     <Option value=\"\" name=\"name\" type=\"QString\"/>     <Option name=\"properties\"/>     <Option value=\"collection\" name=\"type\" type=\"QString\"/>    </Option>   </data-defined-properties>  </renderer-v2>  <selection mode=\"Default\">   <selectionColor invalid=\"1\"/>   <selectionSymbol>    <symbol frame_rate=\"10\" is_animated=\"0\" name=\"\" force_rhr=\"0\" type=\"line\" clip_to_extent=\"1\" alpha=\"1\">     <data_defined_properties>      <Option type=\"Map\">       <Option value=\"\" name=\"name\" type=\"QString\"/>       <Option name=\"properties\"/>       <Option value=\"collection\" name=\"type\" type=\"QString\"/>      </Option>     </data_defined_properties>     <layer pass=\"0\" class=\"SimpleLine\" id=\"{de8ac958-4167-4765-b8e7-9ceedea1870a}\" locked=\"0\" enabled=\"1\">      <Option type=\"Map\">       <Option value=\"0\" name=\"align_dash_pattern\" type=\"QString\"/>       <Option value=\"square\" name=\"capstyle\" type=\"QString\"/>       <Option value=\"5;2\" name=\"customdash\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"customdash_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"customdash_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"dash_pattern_offset\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"dash_pattern_offset_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"dash_pattern_offset_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"draw_inside_polygon\" type=\"QString\"/>       <Option value=\"bevel\" name=\"joinstyle\" type=\"QString\"/>       <Option value=\"35,35,35,255,rgb:0.1372549,0.1372549,0.1372549,1\" name=\"line_color\" type=\"QString\"/>       <Option value=\"solid\" name=\"line_style\" type=\"QString\"/>       <Option value=\"0.26\" name=\"line_width\" type=\"QString\"/>       <Option value=\"MM\" name=\"line_width_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"offset\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"offset_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"offset_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"ring_filter\" type=\"QString\"/>       <Option value=\"0\" name=\"trim_distance_end\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"trim_distance_end_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"trim_distance_end_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"trim_distance_start\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"trim_distance_start_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"trim_distance_start_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"tweak_dash_pattern_on_corners\" type=\"QString\"/>       <Option value=\"0\" name=\"use_custom_dash\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"width_map_unit_scale\" type=\"QString\"/>      </Option>      <data_defined_properties>       <Option type=\"Map\">        <Option value=\"\" name=\"name\" type=\"QString\"/>        <Option name=\"properties\"/>        <Option value=\"collection\" name=\"type\" type=\"QString\"/>       </Option>      </data_defined_properties>     </layer>    </symbol>   </selectionSymbol>  </selection>  <blendMode>0</blendMode>  <featureBlendMode>0</featureBlendMode>  <layerGeometryType>1</layerGeometryType> </qgis>', '<?xml version= \"1.0 \" encoding= \"UTF-8 \"?> <StyledLayerDescriptor xmlns= \"http://www.opengis.net/sld \" xsi:schemaLocation= \"http://www.opengis.net/sld http://schemas.opengis.net/sld/1.1.0/StyledLayerDescriptor.xsd \" xmlns:ogc= \"http://www.opengis.net/ogc \" xmlns:xsi= \"http://www.w3.org/2001/XMLSchema-instance \" xmlns:xlink= \"http://www.w3.org/1999/xlink \" xmlns:se= \"http://www.opengis.net/se \" version= \"1.1.0 \">  <NamedLayer>   <se:Name>routes_cancelled</se:Name>   <UserStyle>    <se:Name>routes_cancelled</se:Name>    <se:FeatureTypeStyle>     <se:Rule>      <se:Name>Single symbol</se:Name>      <se:LineSymbolizer>       <se:Stroke>        <se:SvgParameter name= \"stroke \">#db1e2a</se:SvgParameter>        <se:SvgParameter name= \"stroke-width \">2</se:SvgParameter>        <se:SvgParameter name= \"stroke-linejoin \">bevel</se:SvgParameter>        <se:SvgParameter name= \"stroke-linecap \">square</se:SvgParameter>       </se:Stroke>      </se:LineSymbolizer>     </se:Rule>    </se:FeatureTypeStyle>   </UserStyle>  </NamedLayer> </StyledLayerDescriptor>', 0, 'Mi. Dez. 10 12:22:15 2025', '', null, '2025-12-10T11:22:15.000Z');")

        con.commit()
        con.close()

        wb.save(excel_filename)

    def _create_route_from_edge_ids(self, context: Context, routes: Iterable[str]) -> MultiLineString:
        # get geometries from edge IDs
        geoms = context.routes.es.select(name_in=routes)['geom']
        geom = force_2d(union_all(geoms))
        if geom.is_empty:
            logger.warning(f"Empty route for endpoint {endpoint}")
            exit(9)
        if geom.geom_type == 'LineString':
            return MultiLineString([geom.coords])
        return geom

    def _calculate_totals(self, context: Context):
        filename = os.path.join(self.folder, "route_totals.gpkg")
        out = fiona.open(filename, 'w', layer='routes', driver='GPKG', crs='EPSG:4326', schema={'geometry': 'LineString', 'properties': {'id': 'str', 'start_hub': 'str', 'end_hub': 'str', 'type': 'str', 'attempted': 'int', 'succeeded': 'int'}})

        # create route entries for each route
        for e in context.routes.es:
            attempted = 0
            succeeded = 0

            if e['name'] in self.counters:
                attempted = self.counters[e['name']]['attempted']
                succeeded = self.counters[e['name']]['succeeded']

            out.write({'geometry': force_2d(e['geom']), 'properties': {
                'id': e['name'],
                'start_hub': e.source_vertex['name'],
                'end_hub': e.target_vertex['name'],
                'type': e['type'],
                'attempted': attempted,
                'succeeded': succeeded,
            }})

        out = fiona.open(filename, 'w', layer='hubs', driver='GPKG', crs='EPSG:4326', schema={'geometry': 'Point', 'properties': {'id': 'str'}})

        # create hub entries for each hub
        for e in context.routes.vs:
            out.write({'geometry': {'type': 'Point', 'coordinates': (e['geom'].x, e['geom'].y)}, 'properties': {'id': e['name']}})

        out.close()

        # copy styles for QGIS
        con = sqlite3.connect(filename)
        cur = con.cursor()
        cur.execute("CREATE TABLE layer_styles (id INTEGER PRIMARY KEY, f_table_catalog TEXT, f_table_schema TEXT, f_table_name TEXT, f_geometry_column TEXT, styleName TEXT, styleQML TEXT, styleSLD TEXT, useAsDefault BOOLEAN, description TEXT, owner TEXT, ui TEXT, update_time DATETIME)")
        con.commit()
        cur.execute("INSERT INTO layer_styles (id, f_table_catalog, f_table_schema, f_table_name, f_geometry_column, styleName, styleQML, styleSLD, useAsDefault, description, owner, ui, update_time) VALUES (1, '', '', 'routes', 'geom', 'Routes', '<!DOCTYPE qgis PUBLIC ''http://mrcc.com/qgis.dtd'' ''SYSTEM''> <qgis version=\"3.44.5-Solothurn\" styleCategories=\"Symbology\">  <renderer-v2 symbollevels=\"0\" forceraster=\"0\" referencescale=\"-1\" type=\"singleSymbol\" enableorderby=\"0\">   <symbols>    <symbol clip_to_extent=\"1\" is_animated=\"0\" frame_rate=\"10\" alpha=\"1\" name=\"0\" type=\"line\" force_rhr=\"0\">     <data_defined_properties>      <Option type=\"Map\">       <Option value=\"\" name=\"name\" type=\"QString\"/>       <Option name=\"properties\"/>       <Option value=\"collection\" name=\"type\" type=\"QString\"/>      </Option>     </data_defined_properties>     <layer locked=\"0\" enabled=\"1\" id=\"{9fc73e26-f0e0-4909-aa5f-fc6c2069adcc}\" pass=\"0\" class=\"SimpleLine\">      <Option type=\"Map\">       <Option value=\"0\" name=\"align_dash_pattern\" type=\"QString\"/>       <Option value=\"square\" name=\"capstyle\" type=\"QString\"/>       <Option value=\"5;2\" name=\"customdash\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"customdash_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"customdash_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"dash_pattern_offset\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"dash_pattern_offset_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"dash_pattern_offset_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"draw_inside_polygon\" type=\"QString\"/>       <Option value=\"bevel\" name=\"joinstyle\" type=\"QString\"/>       <Option value=\"0,0,0,255,hsv:0.58938888888888885,0.60439459830624853,0,1\" name=\"line_color\" type=\"QString\"/>       <Option value=\"solid\" name=\"line_style\" type=\"QString\"/>       <Option value=\"1.46\" name=\"line_width\" type=\"QString\"/>       <Option value=\"MM\" name=\"line_width_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"offset\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"offset_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"offset_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"ring_filter\" type=\"QString\"/>       <Option value=\"0\" name=\"trim_distance_end\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"trim_distance_end_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"trim_distance_end_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"trim_distance_start\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"trim_distance_start_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"trim_distance_start_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"tweak_dash_pattern_on_corners\" type=\"QString\"/>       <Option value=\"0\" name=\"use_custom_dash\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"width_map_unit_scale\" type=\"QString\"/>      </Option>      <data_defined_properties>       <Option type=\"Map\">        <Option value=\"\" name=\"name\" type=\"QString\"/>        <Option name=\"properties\" type=\"Map\">         <Option name=\"outlineColor\" type=\"Map\">          <Option value=\"true\" name=\"active\" type=\"bool\"/>          <Option value=\"if (succeeded, ''#000000'', ''#ff0000'')\" name=\"expression\" type=\"QString\"/>          <Option value=\"3\" name=\"type\" type=\"int\"/>         </Option>        </Option>        <Option value=\"collection\" name=\"type\" type=\"QString\"/>       </Option>      </data_defined_properties>     </layer>    </symbol>   </symbols>   <rotation/>   <sizescale/>   <data-defined-properties>    <Option type=\"Map\">     <Option value=\"\" name=\"name\" type=\"QString\"/>     <Option name=\"properties\"/>     <Option value=\"collection\" name=\"type\" type=\"QString\"/>    </Option>   </data-defined-properties>  </renderer-v2>  <selection mode=\"Default\">   <selectionColor invalid=\"1\"/>   <selectionSymbol>    <symbol clip_to_extent=\"1\" is_animated=\"0\" frame_rate=\"10\" alpha=\"1\" name=\"\" type=\"line\" force_rhr=\"0\">     <data_defined_properties>      <Option type=\"Map\">       <Option value=\"\" name=\"name\" type=\"QString\"/>       <Option name=\"properties\"/>       <Option value=\"collection\" name=\"type\" type=\"QString\"/>      </Option>     </data_defined_properties>     <layer locked=\"0\" enabled=\"1\" id=\"{b1ab1734-e5d7-4c63-99ad-62d6b7742248}\" pass=\"0\" class=\"SimpleLine\">      <Option type=\"Map\">       <Option value=\"0\" name=\"align_dash_pattern\" type=\"QString\"/>       <Option value=\"square\" name=\"capstyle\" type=\"QString\"/>       <Option value=\"5;2\" name=\"customdash\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"customdash_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"customdash_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"dash_pattern_offset\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"dash_pattern_offset_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"dash_pattern_offset_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"draw_inside_polygon\" type=\"QString\"/>       <Option value=\"bevel\" name=\"joinstyle\" type=\"QString\"/>       <Option value=\"35,35,35,255,rgb:0.1372549,0.1372549,0.1372549,1\" name=\"line_color\" type=\"QString\"/>       <Option value=\"solid\" name=\"line_style\" type=\"QString\"/>       <Option value=\"0.26\" name=\"line_width\" type=\"QString\"/>       <Option value=\"MM\" name=\"line_width_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"offset\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"offset_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"offset_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"ring_filter\" type=\"QString\"/>       <Option value=\"0\" name=\"trim_distance_end\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"trim_distance_end_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"trim_distance_end_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"trim_distance_start\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"trim_distance_start_map_unit_scale\" type=\"QString\"/>       <Option value=\"MM\" name=\"trim_distance_start_unit\" type=\"QString\"/>       <Option value=\"0\" name=\"tweak_dash_pattern_on_corners\" type=\"QString\"/>       <Option value=\"0\" name=\"use_custom_dash\" type=\"QString\"/>       <Option value=\"3x:0,0,0,0,0,0\" name=\"width_map_unit_scale\" type=\"QString\"/>      </Option>      <data_defined_properties>       <Option type=\"Map\">        <Option value=\"\" name=\"name\" type=\"QString\"/>        <Option name=\"properties\"/>        <Option value=\"collection\" name=\"type\" type=\"QString\"/>       </Option>      </data_defined_properties>     </layer>    </symbol>   </selectionSymbol>  </selection>  <blendMode>0</blendMode>  <featureBlendMode>0</featureBlendMode>  <layerGeometryType>1</layerGeometryType> </qgis> ', '<?xml version=\"1.0\" encoding=\"UTF-8\"?> <StyledLayerDescriptor xmlns=\"http://www.opengis.net/sld\" xmlns:ogc=\"http://www.opengis.net/ogc\" xmlns:se=\"http://www.opengis.net/se\" xmlns:xlink=\"http://www.w3.org/1999/xlink\" xsi:schemaLocation=\"http://www.opengis.net/sld http://schemas.opengis.net/sld/1.1.0/StyledLayerDescriptor.xsd\" version=\"1.1.0\" xmlns:xsi=\"http://www.w3.org/2001/XMLSchema-instance\">  <NamedLayer>   <se:Name>routes</se:Name>   <UserStyle>    <se:Name>routes</se:Name>    <se:FeatureTypeStyle>     <se:Rule>      <se:Name>Single symbol</se:Name>      <se:LineSymbolizer>       <se:Stroke>        <se:SvgParameter name=\"stroke\">#000000</se:SvgParameter>        <se:SvgParameter name=\"stroke-width\">5</se:SvgParameter>        <se:SvgParameter name=\"stroke-linejoin\">bevel</se:SvgParameter>        <se:SvgParameter name=\"stroke-linecap\">square</se:SvgParameter>       </se:Stroke>      </se:LineSymbolizer>     </se:Rule>    </se:FeatureTypeStyle>   </UserStyle>  </NamedLayer> </StyledLayerDescriptor> ', 0, 'Mo. Dez. 8 15:55:36 2025', '', null, '2025-12-08T14:43:24Z');")
        cur.execute("INSERT INTO layer_styles (id, f_table_catalog, f_table_schema, f_table_name, f_geometry_column, styleName, styleQML, styleSLD, useAsDefault, description, owner, ui, update_time) VALUES (2, '', '', 'hubs', 'geom', 'Hubs', '<!DOCTYPE qgis PUBLIC ''http://mrcc.com/qgis.dtd'' ''SYSTEM''> <qgis version=\"3.44.5-Solothurn\" styleCategories=\"Symbology\">  <renderer-v2 type=\"singleSymbol\" forceraster=\"0\" enableorderby=\"0\" symbollevels=\"0\" referencescale=\"-1\">   <symbols>    <symbol type=\"marker\" force_rhr=\"0\" frame_rate=\"10\" alpha=\"1\" is_animated=\"0\" name=\"0\" clip_to_extent=\"1\">     <data_defined_properties>      <Option type=\"Map\">       <Option type=\"QString\" name=\"name\" value=\"\"/>       <Option name=\"properties\"/>       <Option type=\"QString\" name=\"type\" value=\"collection\"/>      </Option>     </data_defined_properties>     <layer class=\"SimpleMarker\" locked=\"0\" pass=\"0\" id=\"{cfd366ce-b938-4a2f-b90e-b90b9d2b5611}\" enabled=\"1\">      <Option type=\"Map\">       <Option type=\"QString\" name=\"angle\" value=\"0\"/>       <Option type=\"QString\" name=\"cap_style\" value=\"square\"/>       <Option type=\"QString\" name=\"color\" value=\"219,30,42,255,rgb:0.8588235,0.1176471,0.1647059,1\"/>       <Option type=\"QString\" name=\"horizontal_anchor_point\" value=\"1\"/>       <Option type=\"QString\" name=\"joinstyle\" value=\"bevel\"/>       <Option type=\"QString\" name=\"name\" value=\"circle\"/>       <Option type=\"QString\" name=\"offset\" value=\"0,0\"/>       <Option type=\"QString\" name=\"offset_map_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>       <Option type=\"QString\" name=\"offset_unit\" value=\"MM\"/>       <Option type=\"QString\" name=\"outline_color\" value=\"128,17,25,255,rgb:0.5019608,0.0666667,0.0980392,1\"/>       <Option type=\"QString\" name=\"outline_style\" value=\"solid\"/>       <Option type=\"QString\" name=\"outline_width\" value=\"0.4\"/>       <Option type=\"QString\" name=\"outline_width_map_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>       <Option type=\"QString\" name=\"outline_width_unit\" value=\"MM\"/>       <Option type=\"QString\" name=\"scale_method\" value=\"diameter\"/>       <Option type=\"QString\" name=\"size\" value=\"4\"/>       <Option type=\"QString\" name=\"size_map_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>       <Option type=\"QString\" name=\"size_unit\" value=\"MM\"/>       <Option type=\"QString\" name=\"vertical_anchor_point\" value=\"1\"/>      </Option>      <data_defined_properties>       <Option type=\"Map\">        <Option type=\"QString\" name=\"name\" value=\"\"/>        <Option name=\"properties\"/>        <Option type=\"QString\" name=\"type\" value=\"collection\"/>       </Option>      </data_defined_properties>     </layer>    </symbol>   </symbols>   <rotation/>   <sizescale/>   <data-defined-properties>    <Option type=\"Map\">     <Option type=\"QString\" name=\"name\" value=\"\"/>     <Option name=\"properties\"/>     <Option type=\"QString\" name=\"type\" value=\"collection\"/>    </Option>   </data-defined-properties>  </renderer-v2>  <selection mode=\"Default\">   <selectionColor invalid=\"1\"/>   <selectionSymbol>    <symbol type=\"marker\" force_rhr=\"0\" frame_rate=\"10\" alpha=\"1\" is_animated=\"0\" name=\"\" clip_to_extent=\"1\">     <data_defined_properties>      <Option type=\"Map\">       <Option type=\"QString\" name=\"name\" value=\"\"/>       <Option name=\"properties\"/>       <Option type=\"QString\" name=\"type\" value=\"collection\"/>      </Option>     </data_defined_properties>     <layer class=\"SimpleMarker\" locked=\"0\" pass=\"0\" id=\"{18d43ab1-e5eb-4441-b036-22182bf5b4b2}\" enabled=\"1\">      <Option type=\"Map\">       <Option type=\"QString\" name=\"angle\" value=\"0\"/>       <Option type=\"QString\" name=\"cap_style\" value=\"square\"/>       <Option type=\"QString\" name=\"color\" value=\"255,0,0,255,rgb:1,0,0,1\"/>       <Option type=\"QString\" name=\"horizontal_anchor_point\" value=\"1\"/>       <Option type=\"QString\" name=\"joinstyle\" value=\"bevel\"/>       <Option type=\"QString\" name=\"name\" value=\"circle\"/>       <Option type=\"QString\" name=\"offset\" value=\"0,0\"/>       <Option type=\"QString\" name=\"offset_map_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>       <Option type=\"QString\" name=\"offset_unit\" value=\"MM\"/>       <Option type=\"QString\" name=\"outline_color\" value=\"35,35,35,255,rgb:0.1372549,0.1372549,0.1372549,1\"/>       <Option type=\"QString\" name=\"outline_style\" value=\"solid\"/>       <Option type=\"QString\" name=\"outline_width\" value=\"0\"/>       <Option type=\"QString\" name=\"outline_width_map_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>       <Option type=\"QString\" name=\"outline_width_unit\" value=\"MM\"/>       <Option type=\"QString\" name=\"scale_method\" value=\"diameter\"/>       <Option type=\"QString\" name=\"size\" value=\"2\"/>       <Option type=\"QString\" name=\"size_map_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>       <Option type=\"QString\" name=\"size_unit\" value=\"MM\"/>       <Option type=\"QString\" name=\"vertical_anchor_point\" value=\"1\"/>      </Option>      <data_defined_properties>       <Option type=\"Map\">        <Option type=\"QString\" name=\"name\" value=\"\"/>        <Option name=\"properties\"/>        <Option type=\"QString\" name=\"type\" value=\"collection\"/>       </Option>      </data_defined_properties>     </layer>    </symbol>   </selectionSymbol>  </selection>  <blendMode>0</blendMode>  <featureBlendMode>0</featureBlendMode>  <layerGeometryType>0</layerGeometryType> </qgis> ', '<?xml version=\"1.0\" encoding=\"UTF-8\"?> <StyledLayerDescriptor xmlns=\"http://www.opengis.net/sld\" xmlns:se=\"http://www.opengis.net/se\" xmlns:xsi=\"http://www.w3.org/2001/XMLSchema-instance\" version=\"1.1.0\" xmlns:ogc=\"http://www.opengis.net/ogc\" xsi:schemaLocation=\"http://www.opengis.net/sld http://schemas.opengis.net/sld/1.1.0/StyledLayerDescriptor.xsd\" xmlns:xlink=\"http://www.w3.org/1999/xlink\">  <NamedLayer>   <se:Name>routes — hubs</se:Name>   <UserStyle>    <se:Name>routes — hubs</se:Name>    <se:FeatureTypeStyle>     <se:Rule>      <se:Name>Single symbol</se:Name>      <se:PointSymbolizer>       <se:Graphic>        <se:Mark>         <se:WellKnownName>circle</se:WellKnownName>         <se:Fill>          <se:SvgParameter name=\"fill\">#db1e2a</se:SvgParameter>         </se:Fill>         <se:Stroke>          <se:SvgParameter name=\"stroke\">#801119</se:SvgParameter>          <se:SvgParameter name=\"stroke-width\">1</se:SvgParameter>         </se:Stroke>        </se:Mark>        <se:Size>14</se:Size>       </se:Graphic>      </se:PointSymbolizer>     </se:Rule>    </se:FeatureTypeStyle>   </UserStyle>  </NamedLayer> </StyledLayerDescriptor> ', 0, 'Mo. Dez. 8 15:53:38 2025', '', null, '2025-12-08T14:44:19Z');")

        con.commit()
        con.close()

    def _persist_agents(self, agents: list[Agent], context: Context, current_day: int):
        day_with_zeroes = str(current_day).rjust(3, '0')
        filename = os.path.join(self.folder, f"day_{day_with_zeroes}.gpkg")

        agent_data = fiona.open(filename, 'w', driver='GPKG', layer='agents', crs='EPSG:4326', schema={'geometry': 'LineString', 'properties': {'id': 'str', 'start_hub': 'str', 'end_hub': 'str', 'start_time': 'datetime', 'end_time': 'datetime', 'is_finished': 'bool', 'is_cancelled': 'bool', 'cancel_reason':'str', 'stops': 'str', 'hubs': 'str', 'edges': 'str'}})

        if self.only_unique:
            self.agent_hashes = set()

        self.start_hubs: dict[str, list[str]] = {}
        self.end_hubs: dict[str, list[str]] = {}

        for agent in agents:
            self._persist_agent(agent, context, agent_data, current_day)

        agent_data.close()

        hub_data = fiona.open(filename, 'w', driver='GPKG', layer='hubs', crs='EPSG:4326', schema={'geometry': 'Point', 'properties': {'id': 'str', 'is_start': 'bool', 'is_end': 'bool', 'is_both': 'bool', 'start_agents': 'str', 'end_agents': 'str'}})

        for hub_id in self.start_hubs:
            hub = context.routes.vs.find(name=hub_id)
            # test end hubs, too
            is_end = False
            end_agents = ''
            if hub_id in self.end_hubs:
                is_end = True
                end_agents = '\n'.join(self.end_hubs[hub_id])

            hub_data.write({'geometry': force_2d(hub['geom']), 'properties': {
                'id': hub_id,
                'is_start': True,
                'is_end': is_end,
                'is_both': is_end,
                'start_agents': '\n'.join(self.start_hubs[hub_id]),
                'end_agents': end_agents,
            }})

        for hub_id in self.end_hubs:
            # if we were in start hubs, skip
            if hub_id in self.start_hubs:
                continue

            hub = context.routes.vs.find(name=hub_id)
            hub_data.write({'geometry': force_2d(hub['geom']), 'properties': {
                'id': hub_id,
                'is_start': False,
                'is_end': True,
                'is_both': False,
                'start_agents': '',
                'end_agents': '\n'.join(self.end_hubs[hub_id]),
            }})

        hub_data.close()

        # copy styles for QGIS
        con = sqlite3.connect(filename)
        cur = con.cursor()
        cur.execute("CREATE TABLE layer_styles (id INTEGER PRIMARY KEY, f_table_catalog TEXT, f_table_schema TEXT, f_table_name TEXT, f_geometry_column TEXT, styleName TEXT, styleQML TEXT, styleSLD TEXT, useAsDefault BOOLEAN, description TEXT, owner TEXT, ui TEXT, update_time DATETIME)")
        con.commit()
        cur.execute("INSERT INTO layer_styles (id, f_table_catalog, f_table_schema, f_table_name, f_geometry_column, styleName, styleQML, styleSLD, useAsDefault, description, owner, ui, update_time) VALUES (1, '', '', 'agents', 'geom', 'Routes', '<!DOCTYPE qgis PUBLIC ''http://mrcc.com/qgis.dtd'' ''SYSTEM''> <qgis version=\"3.44.5-Solothurn\" styleCategories=\"Symbology\">  <renderer-v2 type=\"singleSymbol\" enableorderby=\"0\" symbollevels=\"0\" forceraster=\"0\" referencescale=\"-1\">   <symbols>    <symbol alpha=\"1\" type=\"line\" frame_rate=\"10\" name=\"0\" force_rhr=\"0\" is_animated=\"0\" clip_to_extent=\"1\">     <data_defined_properties>      <Option type=\"Map\">       <Option type=\"QString\" name=\"name\" value=\"\"/>       <Option name=\"properties\"/>       <Option type=\"QString\" name=\"type\" value=\"collection\"/>      </Option>     </data_defined_properties>     <layer id=\"{3a5c8149-3214-4108-b5b4-17266b4549b7}\" class=\"ArrowLine\" pass=\"0\" enabled=\"1\" locked=\"0\">      <Option type=\"Map\">       <Option type=\"QString\" name=\"arrow_start_width\" value=\"0.4\"/>       <Option type=\"QString\" name=\"arrow_start_width_unit\" value=\"MM\"/>       <Option type=\"QString\" name=\"arrow_start_width_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>       <Option type=\"QString\" name=\"arrow_type\" value=\"0\"/>       <Option type=\"QString\" name=\"arrow_width\" value=\"0.4\"/>       <Option type=\"QString\" name=\"arrow_width_unit\" value=\"MM\"/>       <Option type=\"QString\" name=\"arrow_width_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>       <Option type=\"QString\" name=\"head_length\" value=\"0.9\"/>       <Option type=\"QString\" name=\"head_length_unit\" value=\"MM\"/>       <Option type=\"QString\" name=\"head_length_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>       <Option type=\"QString\" name=\"head_thickness\" value=\"0.9\"/>       <Option type=\"QString\" name=\"head_thickness_unit\" value=\"MM\"/>       <Option type=\"QString\" name=\"head_thickness_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>       <Option type=\"QString\" name=\"head_type\" value=\"0\"/>       <Option type=\"QString\" name=\"is_curved\" value=\"0\"/>       <Option type=\"QString\" name=\"is_repeated\" value=\"1\"/>       <Option type=\"QString\" name=\"offset\" value=\"0\"/>       <Option type=\"QString\" name=\"offset_unit\" value=\"MM\"/>       <Option type=\"QString\" name=\"offset_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>       <Option type=\"QString\" name=\"ring_filter\" value=\"0\"/>      </Option>      <data_defined_properties>       <Option type=\"Map\">        <Option type=\"QString\" name=\"name\" value=\"\"/>        <Option name=\"properties\"/>        <Option type=\"QString\" name=\"type\" value=\"collection\"/>       </Option>      </data_defined_properties>      <symbol alpha=\"1\" type=\"fill\" frame_rate=\"10\" name=\"@0@0\" force_rhr=\"0\" is_animated=\"0\" clip_to_extent=\"1\">       <data_defined_properties>        <Option type=\"Map\">         <Option type=\"QString\" name=\"name\" value=\"\"/>         <Option name=\"properties\"/>         <Option type=\"QString\" name=\"type\" value=\"collection\"/>        </Option>       </data_defined_properties>       <layer id=\"{35fcbb28-dc59-4981-880e-889fe2eadaca}\" class=\"SimpleFill\" pass=\"0\" enabled=\"1\" locked=\"0\">        <Option type=\"Map\">         <Option type=\"QString\" name=\"border_width_map_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>         <Option type=\"QString\" name=\"color\" value=\"255,0,0,255,hsv:0,1,1,1\"/>         <Option type=\"QString\" name=\"joinstyle\" value=\"bevel\"/>         <Option type=\"QString\" name=\"offset\" value=\"0,0\"/>         <Option type=\"QString\" name=\"offset_map_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>         <Option type=\"QString\" name=\"offset_unit\" value=\"MM\"/>         <Option type=\"QString\" name=\"outline_color\" value=\"219,30,42,255,rgb:0.8588235,0.1176471,0.1647059,1\"/>         <Option type=\"QString\" name=\"outline_style\" value=\"no\"/>         <Option type=\"QString\" name=\"outline_width\" value=\"0.66\"/>         <Option type=\"QString\" name=\"outline_width_unit\" value=\"MM\"/>         <Option type=\"QString\" name=\"style\" value=\"solid\"/>        </Option>        <data_defined_properties>         <Option type=\"Map\">          <Option type=\"QString\" name=\"name\" value=\"\"/>          <Option name=\"properties\"/>          <Option type=\"QString\" name=\"type\" value=\"collection\"/>         </Option>        </data_defined_properties>       </layer>      </symbol>     </layer>    </symbol>   </symbols>   <rotation/>   <sizescale/>   <data-defined-properties>    <Option type=\"Map\">     <Option type=\"QString\" name=\"name\" value=\"\"/>     <Option name=\"properties\"/>     <Option type=\"QString\" name=\"type\" value=\"collection\"/>    </Option>   </data-defined-properties>  </renderer-v2>  <selection mode=\"Default\">   <selectionColor invalid=\"1\"/>   <selectionSymbol>    <symbol alpha=\"1\" type=\"line\" frame_rate=\"10\" name=\"\" force_rhr=\"0\" is_animated=\"0\" clip_to_extent=\"1\">     <data_defined_properties>      <Option type=\"Map\">       <Option type=\"QString\" name=\"name\" value=\"\"/>       <Option name=\"properties\"/>       <Option type=\"QString\" name=\"type\" value=\"collection\"/>      </Option>     </data_defined_properties>     <layer id=\"{4cddcb78-295d-426c-9aef-2e48bf5b1d55}\" class=\"SimpleLine\" pass=\"0\" enabled=\"1\" locked=\"0\">      <Option type=\"Map\">       <Option type=\"QString\" name=\"align_dash_pattern\" value=\"0\"/>       <Option type=\"QString\" name=\"capstyle\" value=\"square\"/>       <Option type=\"QString\" name=\"customdash\" value=\"5;2\"/>       <Option type=\"QString\" name=\"customdash_map_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>       <Option type=\"QString\" name=\"customdash_unit\" value=\"MM\"/>       <Option type=\"QString\" name=\"dash_pattern_offset\" value=\"0\"/>       <Option type=\"QString\" name=\"dash_pattern_offset_map_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>       <Option type=\"QString\" name=\"dash_pattern_offset_unit\" value=\"MM\"/>       <Option type=\"QString\" name=\"draw_inside_polygon\" value=\"0\"/>       <Option type=\"QString\" name=\"joinstyle\" value=\"bevel\"/>       <Option type=\"QString\" name=\"line_color\" value=\"35,35,35,255,rgb:0.1372549,0.1372549,0.1372549,1\"/>       <Option type=\"QString\" name=\"line_style\" value=\"solid\"/>       <Option type=\"QString\" name=\"line_width\" value=\"0.26\"/>       <Option type=\"QString\" name=\"line_width_unit\" value=\"MM\"/>       <Option type=\"QString\" name=\"offset\" value=\"0\"/>       <Option type=\"QString\" name=\"offset_map_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>       <Option type=\"QString\" name=\"offset_unit\" value=\"MM\"/>       <Option type=\"QString\" name=\"ring_filter\" value=\"0\"/>       <Option type=\"QString\" name=\"trim_distance_end\" value=\"0\"/>       <Option type=\"QString\" name=\"trim_distance_end_map_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>       <Option type=\"QString\" name=\"trim_distance_end_unit\" value=\"MM\"/>       <Option type=\"QString\" name=\"trim_distance_start\" value=\"0\"/>       <Option type=\"QString\" name=\"trim_distance_start_map_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>       <Option type=\"QString\" name=\"trim_distance_start_unit\" value=\"MM\"/>       <Option type=\"QString\" name=\"tweak_dash_pattern_on_corners\" value=\"0\"/>       <Option type=\"QString\" name=\"use_custom_dash\" value=\"0\"/>       <Option type=\"QString\" name=\"width_map_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>      </Option>      <data_defined_properties>       <Option type=\"Map\">        <Option type=\"QString\" name=\"name\" value=\"\"/>        <Option name=\"properties\"/>        <Option type=\"QString\" name=\"type\" value=\"collection\"/>       </Option>      </data_defined_properties>     </layer>    </symbol>   </selectionSymbol>  </selection>  <blendMode>0</blendMode>  <featureBlendMode>0</featureBlendMode>  <layerGeometryType>1</layerGeometryType> </qgis> ', '', 0, 'Mo. Dez. 8 15:01:42 2025', '', null, '2025-12-08T13:46:29Z');")
        cur.execute("INSERT INTO layer_styles (id, f_table_catalog, f_table_schema, f_table_name, f_geometry_column, styleName, styleQML, styleSLD, useAsDefault, description, owner, ui, update_time) VALUES (2, '', '', 'hubs', 'geom', 'Hubs', '<!DOCTYPE qgis PUBLIC ''http://mrcc.com/qgis.dtd'' ''SYSTEM''> <qgis styleCategories=\"Symbology\" version=\"3.44.5-Solothurn\">  <renderer-v2 type=\"singleSymbol\" referencescale=\"-1\" symbollevels=\"0\" enableorderby=\"0\" forceraster=\"0\">   <symbols>    <symbol type=\"marker\" alpha=\"1\" frame_rate=\"10\" name=\"0\" force_rhr=\"0\" clip_to_extent=\"1\" is_animated=\"0\">     <data_defined_properties>      <Option type=\"Map\">       <Option type=\"QString\" name=\"name\" value=\"\"/>       <Option name=\"properties\"/>       <Option type=\"QString\" name=\"type\" value=\"collection\"/>      </Option>     </data_defined_properties>     <layer pass=\"0\" enabled=\"1\" class=\"SimpleMarker\" id=\"{f1df8086-593d-4e06-a17f-a5d339956895}\" locked=\"0\">      <Option type=\"Map\">       <Option type=\"QString\" name=\"angle\" value=\"0\"/>       <Option type=\"QString\" name=\"cap_style\" value=\"square\"/>       <Option type=\"QString\" name=\"color\" value=\"219,30,42,255,rgb:0.8588235,0.1176471,0.1647059,1\"/>       <Option type=\"QString\" name=\"horizontal_anchor_point\" value=\"1\"/>       <Option type=\"QString\" name=\"joinstyle\" value=\"bevel\"/>       <Option type=\"QString\" name=\"name\" value=\"circle\"/>       <Option type=\"QString\" name=\"offset\" value=\"0,0\"/>       <Option type=\"QString\" name=\"offset_map_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>       <Option type=\"QString\" name=\"offset_unit\" value=\"MM\"/>       <Option type=\"QString\" name=\"outline_color\" value=\"128,17,25,255,rgb:0.5019608,0.0666667,0.0980392,1\"/>       <Option type=\"QString\" name=\"outline_style\" value=\"solid\"/>       <Option type=\"QString\" name=\"outline_width\" value=\"0.4\"/>       <Option type=\"QString\" name=\"outline_width_map_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>       <Option type=\"QString\" name=\"outline_width_unit\" value=\"MM\"/>       <Option type=\"QString\" name=\"scale_method\" value=\"diameter\"/>       <Option type=\"QString\" name=\"size\" value=\"4\"/>       <Option type=\"QString\" name=\"size_map_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>       <Option type=\"QString\" name=\"size_unit\" value=\"MM\"/>       <Option type=\"QString\" name=\"vertical_anchor_point\" value=\"1\"/>      </Option>      <data_defined_properties>       <Option type=\"Map\">        <Option type=\"QString\" name=\"name\" value=\"\"/>        <Option type=\"Map\" name=\"properties\">         <Option type=\"Map\" name=\"fillColor\">          <Option type=\"bool\" name=\"active\" value=\"true\"/>          <Option type=\"QString\" name=\"expression\" value=\"if(is_both, ''yellow'',  if (is_start, ''purple'', ''coral''))\"/>          <Option type=\"int\" name=\"type\" value=\"3\"/>         </Option>        </Option>        <Option type=\"QString\" name=\"type\" value=\"collection\"/>       </Option>      </data_defined_properties>     </layer>    </symbol>   </symbols>   <rotation/>   <sizescale/>   <data-defined-properties>    <Option type=\"Map\">     <Option type=\"QString\" name=\"name\" value=\"\"/>     <Option name=\"properties\"/>     <Option type=\"QString\" name=\"type\" value=\"collection\"/>    </Option>   </data-defined-properties>  </renderer-v2>  <selection mode=\"Default\">   <selectionColor invalid=\"1\"/>   <selectionSymbol>    <symbol type=\"marker\" alpha=\"1\" frame_rate=\"10\" name=\"\" force_rhr=\"0\" clip_to_extent=\"1\" is_animated=\"0\">     <data_defined_properties>      <Option type=\"Map\">       <Option type=\"QString\" name=\"name\" value=\"\"/>       <Option name=\"properties\"/>       <Option type=\"QString\" name=\"type\" value=\"collection\"/>      </Option>     </data_defined_properties>     <layer pass=\"0\" enabled=\"1\" class=\"SimpleMarker\" id=\"{2c803217-9df9-441a-af73-cf018fd5d9b5}\" locked=\"0\">      <Option type=\"Map\">       <Option type=\"QString\" name=\"angle\" value=\"0\"/>       <Option type=\"QString\" name=\"cap_style\" value=\"square\"/>       <Option type=\"QString\" name=\"color\" value=\"255,0,0,255,rgb:1,0,0,1\"/>       <Option type=\"QString\" name=\"horizontal_anchor_point\" value=\"1\"/>       <Option type=\"QString\" name=\"joinstyle\" value=\"bevel\"/>       <Option type=\"QString\" name=\"name\" value=\"circle\"/>       <Option type=\"QString\" name=\"offset\" value=\"0,0\"/>       <Option type=\"QString\" name=\"offset_map_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>       <Option type=\"QString\" name=\"offset_unit\" value=\"MM\"/>       <Option type=\"QString\" name=\"outline_color\" value=\"35,35,35,255,rgb:0.1372549,0.1372549,0.1372549,1\"/>       <Option type=\"QString\" name=\"outline_style\" value=\"solid\"/>       <Option type=\"QString\" name=\"outline_width\" value=\"0\"/>       <Option type=\"QString\" name=\"outline_width_map_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>       <Option type=\"QString\" name=\"outline_width_unit\" value=\"MM\"/>       <Option type=\"QString\" name=\"scale_method\" value=\"diameter\"/>       <Option type=\"QString\" name=\"size\" value=\"2\"/>       <Option type=\"QString\" name=\"size_map_unit_scale\" value=\"3x:0,0,0,0,0,0\"/>       <Option type=\"QString\" name=\"size_unit\" value=\"MM\"/>       <Option type=\"QString\" name=\"vertical_anchor_point\" value=\"1\"/>      </Option>      <data_defined_properties>       <Option type=\"Map\">        <Option type=\"QString\" name=\"name\" value=\"\"/>        <Option name=\"properties\"/>        <Option type=\"QString\" name=\"type\" value=\"collection\"/>       </Option>      </data_defined_properties>     </layer>    </symbol>   </selectionSymbol>  </selection>  <blendMode>0</blendMode>  <featureBlendMode>0</featureBlendMode>  <layerGeometryType>0</layerGeometryType> </qgis> ', '<?xml version=\"1.0\" encoding=\"UTF-8\"?> <StyledLayerDescriptor xmlns=\"http://www.opengis.net/sld\" xmlns:se=\"http://www.opengis.net/se\" xmlns:ogc=\"http://www.opengis.net/ogc\" version=\"1.1.0\" xsi:schemaLocation=\"http://www.opengis.net/sld http://schemas.opengis.net/sld/1.1.0/StyledLayerDescriptor.xsd\" xmlns:xsi=\"http://www.w3.org/2001/XMLSchema-instance\" xmlns:xlink=\"http://www.w3.org/1999/xlink\">  <NamedLayer>   <se:Name>day_001 — hubs</se:Name>   <UserStyle>    <se:Name>day_001 — hubs</se:Name>    <se:FeatureTypeStyle>     <se:Rule>      <se:Name>Single symbol</se:Name>      <se:PointSymbolizer>       <se:Graphic>        <se:Mark>         <se:WellKnownName>circle</se:WellKnownName>         <se:Fill>          <se:SvgParameter name=\"fill\">#db1e2a</se:SvgParameter>         </se:Fill>         <se:Stroke>          <se:SvgParameter name=\"stroke\">#801119</se:SvgParameter>          <se:SvgParameter name=\"stroke-width\">1</se:SvgParameter>         </se:Stroke>        </se:Mark>        <se:Size>14</se:Size>       </se:Graphic>      </se:PointSymbolizer>     </se:Rule>    </se:FeatureTypeStyle>   </UserStyle>  </NamedLayer> </StyledLayerDescriptor> ', 0, 'Mo. Dez. 8 15:02:54 2025', '', null, '2025-12-08T13:53:20Z'); ")

        con.commit()
        con.close()

    def _persist_agent(self, agent: Agent, context: Context, agent_data: fiona.Collection, current_day: int):
        # get route/geometry
        route = self._merge_route(agent.route, agent.route_reversed, context)

        # do not save agents with no routes
        if self.ignore_empty_agents and route.is_empty:
            return

        # calculate attempts
        for route_id in list(agent.route[1::2]):
            self._increment_route_counter(route_id)
        for route_id in list(agent.route_before_traceback[1::2]):
            self._increment_route_counter(route_id, is_attempt=True)

        # get start/end time
        start_hub, end_hub, start_delta, end_delta = agent.get_start_end()

        # only save unique routes, if setting is so
        if self.only_unique:
            key = (start_hub, end_hub, start_delta, end_delta)
            if key in self.agent_hashes:
                return
            self.agent_hashes.add(key)

        start_time = self.min_time + dt.timedelta(hours=start_delta)
        end_time = self.min_time + dt.timedelta(hours=end_delta)

        # aggregate start and end hubs and times
        if start_hub not in self.start_hubs:
            self.start_hubs[start_hub] = []
        self.start_hubs[start_hub].append(agent.uid + ': ' + start_time.strftime('%Y-%m-%d %H:%M'))
        if end_hub not in self.end_hubs:
            self.end_hubs[end_hub] = []
        self.end_hubs[end_hub].append(agent.uid + ': ' + start_time.strftime('%Y-%m-%d %H:%M'))

        hubs = ','.join(agent.route[::2])
        edges = ','.join(agent.route[1::2])

        agent_data.write({'geometry': route, 'properties': {
            'id': agent.uid,
            'start_hub': start_hub,
            'end_hub': end_hub,
            'start_time': start_time,
            'end_time': end_time,
            'is_finished': agent.is_finished,
            'is_cancelled': agent.is_cancelled,
            'cancel_reason': agent.cancel_reason,
            'stops': str(agent.rest_history),
            'hubs': hubs,
            'edges': edges,
        }})

        # persist to route graph
        self._save_to_route_graph(agent, start_hub, end_hub, start_time, end_time, current_day)

    def _save_to_route_graph(self, agent: Agent, start_hub: str, end_hub: str, start_time: dt.datetime, end_time: dt.datetime, current_day: int):
        # add route as vertex
        self.route_graph.add_vertex(name=agent.uid, start_time=start_time, end_time=end_time, start_hub=start_hub, end_hub=end_hub, route=agent.route[1::2], is_finished=agent.is_finished, is_cancelled=agent.is_cancelled, cancel_reason=agent.cancel_reason, day=current_day)

        # add edges to parents
        for parent in self.route_graph.vs.select(name_in=agent.parents):
            self.route_graph.add_edge(agent.uid, parent['name'])

    def _merge_route(self, route: list[str], route_reversed: list[bool], context: Context) -> LineString | None:
        coordinates = []

        for idx, route_id in enumerate(route[1::2]):
            # get route
            route = context.routes.es.find(name=route_id)
            # get coordinates
            coords = force_2d(route['geom']).coords
            if route_reversed[idx]:
                coords = reversed(coords)
            # join coordinates
            coords = list(coords)
            if len(coordinates) > 0 and coordinates[-1] == coords[0]:
                # last coordinate is equal to first coordinate, remove it
                coordinates.pop()
            coordinates.extend(coords)

        return force_2d(LineString(coordinates))

    def _increment_route_counter(self, route_id: str, is_attempt = False):
        if route_id not in self.counters:
            self.counters[route_id] = {'attempted': 0,'succeeded': 0}

        self.counters[route_id]['attempted'] += 1
        if not is_attempt:
            self.counters[route_id]['succeeded'] += 1
